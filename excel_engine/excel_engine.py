import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

def create_excel_file(filename, directory, sheet_name="Sheet1"):
    """
    Creates or updates an Excel file with the given filename and specified sheet name in the given directory.

    If the file already exists, a new sheet is added unless a sheet with the same name exists,
    in which case that sheet is overwritten.

    Args:
        filename (str): Name of the Excel file to create (e.g., 'report.xlsx').
        directory (str): Directory path where the Excel file should be created.
        sheet_name (str): Name of the worksheet. Defaults to 'Sheet1'.

    Returns:
        str: Full path to the created or updated Excel file, or None if directory doesn't exist.
    """
    if not os.path.isdir(directory):
        print(f"[ERROR] Directory does not exist: {directory}")
        return None

    # Ensure filename has .xlsx extension
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    file_path = os.path.join(directory, filename)

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            # Remove existing sheet to overwrite it
            std = wb[sheet_name]
            wb.remove(std)
        wb.create_sheet(title=sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    wb.save(file_path)
    return file_path

def style_excel_header(file_path, header_names, sheet_name=None):
    """
    Applies a styled header row to the first row of a specific worksheet in an existing Excel file.

    Args:
        file_path (str): Path to the existing Excel (.xlsx) file.
        header_names (list): List of strings representing column headers.
        sheet_name (str, optional): Name of the worksheet to apply the header to. Defaults to active sheet.

    Returns:
        bool: True if successful, False otherwise.
    """
    if not os.path.exists(file_path):
        print(f"[ERROR] File not found: {file_path}")
        return False

    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(header_names, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        wb.save(file_path)
        return True

    except Exception as e:
        print(f"[ERROR] Failed to apply header: {e}")
        return False

def freeze_excel_header_row(file_path, sheet_name=None):
    """
    Freezes the first row of an existing Excel sheet so the header stays visible when scrolling.

    Args:
        file_path (str): Path to the existing Excel file.
        sheet_name (str, optional): Name of the sheet to freeze the header in. 
                                    If None, the active sheet is used.

    Returns:
        bool: True if successful, False if file/sheet not found.
    """
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        # Freeze the first row
        ws.freeze_panes = ws['A2']

        # Save changes
        wb.save(file_path)
        return True

    except FileNotFoundError:
        print(f"[ERROR] File not found: {file_path}")
        return False
    except KeyError:
        print(f"[ERROR] Sheet '{sheet_name}' not found in file: {file_path}")
        return False
    except Exception as e:
        print(f"[ERROR] {str(e)}")
        return False


def create_excel_m(filename: str, directory: Path, headers: list[str], sheet_name: str = "Sheet1"):
    """
    Creates an Excel file with a styled header and frozen header row in a specified sheet.

    Args:
        filename (str): Desired Excel filename, with or without '.xlsx' extension.
        directory (Path): Target directory to save the Excel file.
        headers (list[str]): List of column headers.
        sheet_name (str): Name of the worksheet to add headers to.

    Returns:
        Path: Full path to the created Excel file, or None if there was an error.
    """
    # Ensure filename has the .xlsx extension
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    # Ensure directory exists
    if not directory.exists():
        print(f"[ERROR] Directory does not exist: {directory}")
        return None

    xlsx_path = directory / filename

    # Create the file
    create_excel_file(filename, directory, sheet_name)

    # Apply headers to the specified sheet
    style_excel_header(xlsx_path, headers, sheet_name)

    # Freeze header row in the specified sheet
    freeze_excel_header_row(xlsx_path, sheet_name)

    return xlsx_path


def write_string_to_excel(file_path, sheet_name, cell, text):
    """
    Writes a string to a specific cell in a specified worksheet of an existing xlsx file.
    Creates the worksheet if it doesn't exist. Only modifies the specified worksheet.

    :param file_path: Path to the Excel file
    :param sheet_name: Name of the sheet to write to (creates if missing)
    :param cell: Cell reference as a string (e.g., 'B4')
    :param text: The string to write
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File '{file_path}' does not exist.")

    # Load the workbook
    wb = load_workbook(file_path)

    # Get or create the specified sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Write the text to the specified cell
    ws[cell] = text

    # Save the workbook
    wb.save(file_path)

def autofit_excel_columns(file_path, sheet_name):
    """
    Adjust the width of each column in the specified sheet to fit its longest cell content,
    emulating Excel's auto-fit behavior.

    Args:
        file_path (str or Path): Path to the Excel file.
        sheet_name (str): The name of the worksheet to adjust.
    """
    wb = load_workbook(file_path)

    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] Sheet '{sheet_name}' not found in workbook.")
        return

    ws = wb[sheet_name]

    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column number (1-based)
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max_length + 2  # Padding for readability
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)
